"""월 마감 시 세무 신고용 엑셀(스냅샷)을 생성한다.

라이브 재계산 모델이 아니라 마감 시점의 확정 데이터를 담는 '리포트 산출물'이므로
집계는 파이썬에서 계산해 값으로 기록한다.
"""

import datetime as dt
import os
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.config import settings
from app.models import Account, ExpenseItem, ExpenseReport
from app.models.enums import EvidenceType, PayMethod

EVIDENCE_LABEL = {
    EvidenceType.tax_invoice: "세금계산서",
    EvidenceType.invoice: "계산서",
    EvidenceType.card: "신용카드",
    EvidenceType.cash_receipt: "현금영수증",
    EvidenceType.simple_receipt: "간이영수증",
    EvidenceType.etc: "기타",
}
PAY_LABEL = {
    PayMethod.corporate_card: "법인카드",
    PayMethod.personal_card: "개인카드",
    PayMethod.cash: "현금",
}

_FONT = "맑은 고딕"
_ACCENT = "1C6B4B"
_WON = "#,##0"
_head_font = Font(name=_FONT, size=10, bold=True, color="FFFFFF")
_body_font = Font(name=_FONT, size=10)
_head_fill = PatternFill("solid", fgColor=_ACCENT)
_thin = Side(style="thin", color="D0D7D0")
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_center = Alignment(horizontal="center", vertical="center")
_right = Alignment(horizontal="right", vertical="center")


def _header(ws, headers, widths, row=1):
    from openpyxl.utils import get_column_letter

    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = _head_font
        c.fill = _head_fill
        c.alignment = _center
        c.border = _border
        ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]


def generate_closing_excel(db: Session, period: str) -> str:
    """해당 귀속월의 확정 경비를 엑셀로 생성하고 저장 경로를 반환한다."""
    stmt = (
        select(ExpenseItem)
        .join(ExpenseReport, ExpenseItem.report_id == ExpenseReport.id)
        .where(ExpenseReport.period == period)
    )
    items = list(db.scalars(stmt))
    accounts = {a.id: a.name for a in db.scalars(select(Account))}

    wb = Workbook()

    # 1) 경비내역
    ws = wb.active
    ws.title = "경비내역"
    headers = ["번호", "사용일자", "부서", "팀", "거래처", "공급가액", "부가세액",
               "합계금액", "계정과목", "증빙유형", "부가세공제", "결제수단", "적요"]
    widths = [5, 12, 12, 12, 18, 12, 11, 12, 14, 12, 10, 10, 26]
    _header(ws, headers, widths)
    for idx, it in enumerate(items, start=1):
        vendor = it.vendor.name if it.vendor else ""
        row = [
            idx,
            it.tx_date.isoformat() if it.tx_date else "",
            it.dept_snapshot or "",
            it.team_snapshot or "",
            vendor,
            it.supply_amount,
            it.vat_amount,
            it.total_amount,
            accounts.get(it.account_id, ""),
            EVIDENCE_LABEL.get(it.evidence_type, ""),
            "공제" if it.vat_deductible else "불공제",
            PAY_LABEL.get(it.pay_method, ""),
            it.memo or "",
        ]
        for ci, v in enumerate(row, start=1):
            c = ws.cell(row=idx + 1, column=ci, value=v)
            c.font = _body_font
            c.border = _border
            if ci in (6, 7, 8):
                c.number_format = _WON
                c.alignment = _right

    # 2) 부서·팀별 집계
    by_team = defaultdict(lambda: [0, 0, 0])  # (dept, team) -> [supply, vat, total]
    by_acct = defaultdict(lambda: [0, 0, 0])
    by_ded = defaultdict(lambda: [0, 0, 0])   # "공제"/"불공제"
    for it in items:
        by_team[(it.dept_snapshot or "-", it.team_snapshot or "-")][0] += it.supply_amount
        by_team[(it.dept_snapshot or "-", it.team_snapshot or "-")][1] += it.vat_amount
        by_team[(it.dept_snapshot or "-", it.team_snapshot or "-")][2] += it.total_amount
        an = accounts.get(it.account_id, "미분류")
        by_acct[an][0] += it.supply_amount
        by_acct[an][1] += it.vat_amount
        by_acct[an][2] += it.total_amount
        key = "공제" if it.vat_deductible else "불공제"
        by_ded[key][0] += it.supply_amount
        by_ded[key][1] += it.vat_amount
        by_ded[key][2] += it.total_amount

    def _agg_sheet(title, key_headers, data: dict):
        s = wb.create_sheet(title)
        heads = key_headers + ["공급가액", "부가세액", "합계금액"]
        widths2 = [14] * len(key_headers) + [13, 12, 13]
        _header(s, heads, widths2)
        r = 2
        for key, vals in data.items():
            keys = key if isinstance(key, tuple) else (key,)
            for ci, kv in enumerate(keys, start=1):
                cc = s.cell(row=r, column=ci, value=kv)
                cc.font = _body_font
                cc.border = _border
            base = len(keys)
            for j, v in enumerate(vals):
                cc = s.cell(row=r, column=base + 1 + j, value=v)
                cc.font = _body_font
                cc.border = _border
                cc.number_format = _WON
                cc.alignment = _right
            r += 1

    _agg_sheet("부서·팀별집계", ["부서", "팀"], by_team)
    _agg_sheet("계정과목별집계", ["계정과목"], by_acct)
    _agg_sheet("부가세정리", ["구분"], by_ded)

    os.makedirs(settings.export_dir, exist_ok=True)
    stamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(settings.export_dir, f"경비마감_{period}_{stamp}.xlsx")
    wb.save(path)
    return path
